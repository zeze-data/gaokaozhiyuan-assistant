# -*- coding: utf-8 -*-
import datetime
import os
import re
from typing import TypedDict, cast

import pandas as pd  # pyright: ignore [reportMissingTypeStubs]
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ==================== 配置区域 ====================
INPUT_DIR = "./input"  # 存放TXT文本的文件夹
OUTPUT_DIR = "./output"  # 导出Excel的文件夹

# 考生信息配置
CANDIDATE_RANK = 14580  # 考生今年位次
# ==================================================

# 定义 TypedDict 类型以满足 basedpyright 强类型要求
RegistryEntry = TypedDict(
    "RegistryEntry",
    {
        "批次": str,
        "院校": str,
        "专业组": str,
        "科目要求": str,
        "专业": str,
        "学费": str,
        "计划数": str,
        "25计划数": str,
        "25最低分": str,
        "25最低分位次": str,
        "25平均分": str,
        "25平均位次": str,
        "24最低分": str,
        "24最低分位次": str,
        "24平均分": str,
        "24平均位次": str,
        "计划与位次趋势": str,
        "COLOR_TAG": str,
        "志愿类别": str,
        "CAT_COLOR_TAG": str,
        "所在地": str,
    },
)


def init_environment() -> None:
    """初始化文件夹"""
    for folder in [INPUT_DIR, OUTPUT_DIR]:
        if not os.path.exists(folder):
            os.makedirs(folder)


def safe_str(val: object) -> str:
    """安全地将值转换为字符串，处理 None 和 NaN"""
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str.lower() in ("nan", "<na>", "nat", ""):
        return ""
    return val_str


def clean_major_name(name: object) -> str:
    """清洗专业名称，去除开头的数字和符号干扰"""
    name_str = safe_str(name)
    if not name_str:
        return ""
    return re.sub(r"^\d+[-－＿\s]*", "", name_str)


def clean_group_name(group_name: object) -> str:
    """提取专业组的数字代码部分，例如 '11701-上海电力(01)' -> '11701'"""
    group_str = safe_str(group_name)
    if not group_str:
        return ""
    match = re.match(r"^(\d+)", group_str)
    if match:
        return match.group(1)
    return group_str


def select_best_by_plan(
    candidates: list[tuple[str, str, str]],
    row_dict: dict[str, object],
    merged_registry: dict[tuple[str, str, str], RegistryEntry],
) -> tuple[str, str, str]:
    """在一组候选专业中，通过对比计划数寻找最接近的匹配项"""
    p_score_str = safe_str(row_dict.get("计划数"))
    p_score = int(p_score_str) if p_score_str.isdigit() else 0

    year = safe_str(row_dict.get("招生年份"))
    best_candidate = candidates[0]
    min_diff = float("inf")

    for cand_key in candidates:
        cand_data = merged_registry[cand_key]
        if year == "2024":
            p_ref_str = cand_data["25计划数"] or cand_data["计划数"]
        else:
            p_ref_str = cand_data["计划数"]
        p_ref = int(p_ref_str) if p_ref_str.isdigit() else 0
        diff = abs(p_score - p_ref)
        if diff < min_diff:
            min_diff = diff
            best_candidate = cand_key

    return best_candidate


def parse_txt_file(file_name: str) -> pd.DataFrame | None:
    """通用且高容错的TXT文本解析引擎"""
    file_path = os.path.join(INPUT_DIR, file_name)
    if not os.path.exists(file_path):
        return None

    with open(file_path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]

    if not lines:
        return None

    # 定位表头行
    header_idx = -1
    for idx, line in enumerate(lines):
        if "招生院校" in line or "院校" in line:
            header_idx = idx
            break

    if header_idx == -1:
        return None

    header_line = lines[header_idx]
    # 智能识别分隔符（优先Tab键，其次空格）
    headers = header_line.split("\t") if "\t" in header_line else header_line.split()
    headers = [h.strip() for h in headers if h.strip()]

    data_rows: list[list[str]] = []
    for line in lines[header_idx + 1 :]:
        # 与表头保持相同的切分逻辑
        row = line.split("\t") if "\t" in header_line else line.split()
        row = [item.strip() for item in row]

        if not row or len(row) == 0:
            continue

        # 高容错机制：长对齐裁剪，短对齐用空字串补齐
        if len(row) < len(headers):
            row += [""] * (len(headers) - len(row))
        else:
            row = row[: len(headers)]
        data_rows.append(row)

    return pd.DataFrame(data_rows, columns=headers)


def execute_rolling_backup(output_dir: str, base_name: str) -> None:
    """智能三级滚动备份机制 (-01永远最新)"""
    target_file = os.path.join(output_dir, f"{base_name}.xlsx")
    if not os.path.exists(target_file):
        return  # 没有历史文件，无需备份

    bak1 = os.path.join(output_dir, f"{base_name}-01.xlsx")
    bak2 = os.path.join(output_dir, f"{base_name}-02.xlsx")
    bak3 = os.path.join(output_dir, f"{base_name}-03.xlsx")

    # 倒序滚动推移
    if os.path.exists(bak2):
        if os.path.exists(bak3):
            os.remove(bak3)
        os.rename(bak2, bak3)

    if os.path.exists(bak1):
        os.rename(bak1, bak2)

    os.rename(target_file, bak1)
    print("[备份成功] 旧文件已自动滚入备份链中。")


def generate_report() -> None:
    init_environment()
    today_str = datetime.datetime.now().strftime("%Y%m%d")
    base_filename = f"志愿填报辅助-{today_str}"

    # 1. 读取数据
    df_plan = parse_txt_file("今年招生计划.txt")
    df_score = parse_txt_file("往年院校录取分数.txt")

    if df_plan is None and df_score is None:
        print(
            "错误：未在输入文件夹下找到有效的 '今年招生计划.txt' 或 '往年院校录取分数.txt'。"
        )
        return

    # 2. 核心合并引擎
    merged_registry: dict[tuple[str, str, str], RegistryEntry] = {}

    # 先处理今年计划
    plan_loaded_count = 0
    if df_plan is not None:
        for _, row in df_plan.iterrows():
            row_dict = cast(dict[str, object], row.to_dict())  # pyright: ignore [reportUnknownMemberType]
            school = safe_str(row_dict.get("招生院校"))
            major_raw = safe_str(row_dict.get("专业"))
            if not school or not major_raw:
                continue

            clean_maj = clean_major_name(major_raw)
            group_raw = safe_str(row_dict.get("专业组"))
            group_code = clean_group_name(group_raw)

            # 联合主键：(招生院校, 专业组数字代码, 专业名称)
            key = (school, group_code, clean_maj)

            merged_registry[key] = {
                "批次": "提前批" if group_code.startswith("0") else "普通批",
                "院校": school,
                "专业组": group_raw,
                "科目要求": safe_str(row_dict.get("科目要求")),
                "专业": major_raw,  # 使用包含前缀代号的原始专业名称，保证与今年招生计划显示一致
                "学费": safe_str(row_dict.get("学费")),
                "计划数": safe_str(row_dict.get("计划数")),
                "25计划数": "",
                "25最低分": "",
                "25最低分位次": "",
                "25平均分": "",
                "25平均位次": "",
                "24最低分": "",
                "24最低分位次": "",
                "24平均分": "",
                "24平均位次": "",
                "计划与位次趋势": "",
                "COLOR_TAG": "",
                "志愿类别": "",
                "CAT_COLOR_TAG": "",
                "所在地": safe_str(row_dict.get("所在地")),
            }
            plan_loaded_count += 1

    print(
        f"[数据加载] 从 '今年招生计划.txt' 中成功加载 {plan_loaded_count} 个专业计划。"
    )

    # 再并入往年分数（仅填充招生计划中已有的专业，实现按计划过滤）
    if df_score is not None:
        for _, row in df_score.iterrows():
            row_dict = cast(dict[str, object], row.to_dict())  # pyright: ignore [reportUnknownMemberType]
            school = safe_str(row_dict.get("招生院校"))
            major_raw = safe_str(row_dict.get("专业"))
            if not school or not major_raw:
                continue

            clean_maj = clean_major_name(major_raw)
            group_raw = safe_str(row_dict.get("专业组"))
            group_code = clean_group_name(group_raw)

            # 联合主键：(招生院校, 专业组数字代码, 专业名称)
            key = (school, group_code, clean_maj)

            # 模糊匹配逻辑
            target_key = key
            if key not in merged_registry:
                # 1. 尝试在同学校、同专业组代码下进行模糊匹配（解决专业改名问题，如工科试验班）
                candidates: list[tuple[str, str, str]] = []
                for reg_key in merged_registry.keys():
                    reg_school, reg_group, reg_maj = reg_key
                    if reg_school == school and reg_group == group_code:
                        # 区分普通专业与中外合作办学专业，防止误匹配
                        if ("中外合作" in clean_maj) != ("中外合作" in reg_maj):
                            continue
                        if clean_maj in reg_maj or reg_maj in clean_maj:
                            candidates.append(reg_key)

                if len(candidates) == 1:
                    target_key = candidates[0]
                elif len(candidates) > 1:
                    target_key = select_best_by_plan(
                        candidates, row_dict, merged_registry
                    )
                else:
                    # 2. 尝试跨专业组匹配（解决专业组代码变动问题，如 12105 变 12104）
                    cross_candidates: list[tuple[str, str, str]] = []
                    score_batch = (
                        "提前批"
                        if "提前" in safe_str(row_dict.get("招生批次"))
                        else "普通批"
                    )
                    for reg_key in merged_registry.keys():
                        reg_school, reg_group, reg_maj = reg_key
                        if reg_school == school:
                            if ("中外合作" in clean_maj) != ("中外合作" in reg_maj):
                                continue
                            if merged_registry[reg_key]["批次"] != score_batch:
                                continue
                            if clean_maj == reg_maj:
                                cross_candidates.append(reg_key)

                    # 如果没有名称完全相同的，寻找名称模糊包含的
                    if not cross_candidates:
                        for reg_key in merged_registry.keys():
                            reg_school, reg_group, reg_maj = reg_key
                            if reg_school == school:
                                if ("中外合作" in clean_maj) != ("中外合作" in reg_maj):
                                    continue
                                if merged_registry[reg_key]["批次"] != score_batch:
                                    continue
                                if clean_maj in reg_maj or reg_maj in clean_maj:
                                    cross_candidates.append(reg_key)

                    if len(cross_candidates) == 1:
                        target_key = cross_candidates[0]
                    elif len(cross_candidates) > 1:
                        # 如果有多个候选，优先选择科目要求一致的；如果科目要求也一致，选择招生计划数最接近的
                        subject_req = safe_str(row_dict.get("科目要求"))
                        subject_matches = [
                            c
                            for c in cross_candidates
                            if merged_registry[c]["科目要求"] == subject_req
                        ]
                        if len(subject_matches) == 1:
                            target_key = subject_matches[0]
                        elif len(subject_matches) > 1:
                            target_key = select_best_by_plan(
                                subject_matches, row_dict, merged_registry
                            )
                        else:
                            target_key = select_best_by_plan(
                                cross_candidates, row_dict, merged_registry
                            )
                    else:
                        # 如果仍完全找不到匹配，跳过此行
                        continue

            # 仅在今年招生计划已包含（或通过模糊匹配/跨组匹配映射到）该专业的情况下进行填充
            if target_key not in merged_registry:
                continue

            data_bucket = merged_registry[target_key]
            # 仅补齐可能缺失的科目要求（院校/专业组/专业名字锁定为今年计划名称，不予覆盖！）
            if not data_bucket["科目要求"]:
                data_bucket["科目要求"] = safe_str(row_dict.get("科目要求"))

            batch_str = safe_str(row_dict.get("招生批次"))
            if batch_str:
                if "提前" in batch_str:
                    data_bucket["批次"] = "提前批"
                elif "普通" in batch_str:
                    data_bucket["批次"] = "普通批"

            year = safe_str(row_dict.get("招生年份"))
            if year == "2025":
                data_bucket["25计划数"] = safe_str(row_dict.get("计划数"))
                data_bucket["25最低分"] = safe_str(row_dict.get("最低分"))
                data_bucket["25最低分位次"] = safe_str(row_dict.get("最低分位次"))
                data_bucket["25平均分"] = safe_str(row_dict.get("平均分"))
                data_bucket["25平均位次"] = safe_str(
                    row_dict.get("平均分位次")
                ) or safe_str(row_dict.get("平均位次"))
            elif year == "2024":
                data_bucket["24最低分"] = safe_str(row_dict.get("最低分"))
                data_bucket["24最低分位次"] = safe_str(row_dict.get("最低分位次"))
                data_bucket["24平均分"] = safe_str(row_dict.get("平均分"))
                data_bucket["24平均位次"] = safe_str(
                    row_dict.get("平均分位次")
                ) or safe_str(row_dict.get("平均位次"))

    # 3. 计算趋势与分析，以及志愿填报分类
    final_rows: list[RegistryEntry] = []
    for _, data in merged_registry.items():
        # 3.1 计划与位次趋势计算
        try:
            p_now = int(data["计划数"]) if str(data["计划数"]).isdigit() else None
            p_25 = int(data["25计划数"]) if str(data["25计划数"]).isdigit() else None
            r_25 = (
                int(data["25最低分位次"])
                if str(data["25最低分位次"]).isdigit()
                else None
            )
            r_24 = (
                int(data["24最低分位次"])
                if str(data["24最低分位次"]).isdigit()
                else None
            )

            if p_now is not None and p_25 is not None and abs(p_now - p_25) < 2:
                trend, color_tag = "计划稳定", "YELLOW"
            elif (
                p_now is not None
                and p_25 is not None
                and p_now > p_25
                and r_25 is not None
                and r_24 is not None
                and r_25 > r_24
            ):
                trend, color_tag = "扩招位次下降", "GREEN"
            elif (
                p_now is not None
                and p_25 is not None
                and p_now < p_25
                and r_25 is not None
                and r_24 is not None
                and r_25 < r_24
            ):
                trend, color_tag = "缩招位次上升", "RED"
            elif p_now is not None and p_25 is not None:
                diff = p_now - p_25
                trend = f"扩招({diff}人)" if diff > 0 else f"缩招({abs(diff)}人)"
                color_tag = "LIGHT_GREEN" if diff > 0 else "LIGHT_RED"
            else:
                trend, color_tag = "数据不足", "GRAY"
        except Exception:
            trend, color_tag = "未能计算", "GRAY"

        data["计划与位次趋势"] = trend
        data["COLOR_TAG"] = color_tag

        # 3.2 志愿类别分类计算 (冲、稳、保、难、垫)
        r_25_str = data["25最低分位次"]
        r_24_str = data["24最低分位次"]

        h_rank: int | None = None
        if r_25_str.isdigit():
            h_rank = int(r_25_str)
        elif r_24_str.isdigit():
            h_rank = int(r_24_str)

        if h_rank is not None:
            # 根据考生位次 CANDIDATE_RANK 划分志愿属性
            if h_rank < 0.80 * CANDIDATE_RANK:
                cat = "难"
                cat_color = "RED"
            elif h_rank < 0.95 * CANDIDATE_RANK:
                cat = "冲"
                cat_color = "YELLOW"
            elif h_rank <= 1.05 * CANDIDATE_RANK:
                cat = "稳"
                cat_color = "LIGHT_GREEN"
            elif h_rank <= 1.20 * CANDIDATE_RANK:
                cat = "保"
                cat_color = "MID_GREEN"
            else:
                cat = "垫"
                cat_color = "BLUE"
        else:
            cat = "数据不足"
            cat_color = "GRAY"

        data["志愿类别"] = cat
        data["CAT_COLOR_TAG"] = cat_color

        final_rows.append(data)

    # 3.3 计算每个专业组 (院校, 专业组) 的最小录取位次
    group_min_ranks: dict[tuple[str, str], int] = {}
    for data in final_rows:
        school = data["院校"]
        group_raw = data["专业组"]
        group_key = (school, group_raw)

        r_25_str = data["25最低分位次"]
        r_24_str = data["24最低分位次"]
        r_25 = int(r_25_str) if r_25_str.isdigit() else None
        r_24 = int(r_24_str) if r_24_str.isdigit() else None
        h_rank = r_25 if r_25 is not None else r_24

        if h_rank is not None:
            if group_key not in group_min_ranks or h_rank < group_min_ranks[group_key]:
                group_min_ranks[group_key] = h_rank

    # 3.4 多级自定义排序
    def get_sort_key(data: RegistryEntry) -> tuple[int, int, int, str, str, int]:
        # 1. 批次：提前批在前 (0)，普通批在后 (1)
        batch_order = 0 if data["批次"] == "提前批" else 1

        # 2. 院校专业组难度：组内最难专业的最低分位次（越难位次越小，升序）
        school = data["院校"]
        group_raw = data["专业组"]
        group_key = (school, group_raw)
        group_min_rank = group_min_ranks.get(group_key, 999999)

        # 3. 地域：上海在前 (0)，非上海在后 (1)
        group_is_shanghai = 0 if data["所在地"] == "上海" else 1

        # 4. 组内专业难度（越难位次越小，升序）
        r_25_str = data["25最低分位次"]
        r_24_str = data["24最低分位次"]
        r_25 = int(r_25_str) if r_25_str.isdigit() else None
        r_24 = int(r_24_str) if r_24_str.isdigit() else None
        h_rank = r_25 if r_25 is not None else r_24
        major_rank = h_rank if h_rank is not None else 999999

        return (
            batch_order,
            group_min_rank,
            group_is_shanghai,
            school,
            group_raw,
            major_rank,
        )

    final_rows.sort(key=get_sort_key)

    # 4. 执行备份
    execute_rolling_backup(OUTPUT_DIR, base_filename)

    # 5. 写入高度美化的Excel文件
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("无法获取活动的 Worksheet")

    ws = cast(Worksheet, ws)
    ws.title = "志愿填报辅助看板"
    ws.views.sheetView[0].showGridLines = True  # 强制显示网格线

    headers_order = [
        "批次",
        "院校",
        "专业组",
        "科目要求",
        "专业",
        "志愿类别",  # 将志愿类别放在醒目位置
        "学费",
        "计划数",
        "25计划数",
        "25最低分",
        "25最低分位次",
        "25平均分",
        "25平均位次",
        "24最低分",
        "24最低分位次",
        "24平均分",
        "24平均位次",
        "计划与位次趋势",
    ]

    # 写入表头
    ws.append(headers_order)

    # 样式定义
    header_fill = PatternFill(
        start_color="2F5597", end_color="2F5597", fill_type="solid"
    )
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=10)

    fills = {
        "GREEN": PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        ),  # 趋势用：软绿色
        "YELLOW": PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        ),  # 冲用：软黄色
        "RED": PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        ),  # 难用：软红色
        "LIGHT_GREEN": PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        ),  # 稳用：软浅绿
        "MID_GREEN": PatternFill(
            start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"
        ),  # 保用：软中绿
        "BLUE": PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
        ),  # 垫用：软浅蓝
        "LIGHT_GREEN_TREND": PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        ),
        "LIGHT_RED": PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        ),
        "GRAY": PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        ),
    }

    text_colors = {
        "GREEN": Font(name="微软雅黑", size=10, color="006100", bold=True),
        "YELLOW": Font(name="微软雅黑", size=10, color="9C6500", bold=True),
        "RED": Font(name="微软雅黑", size=10, color="9C0006", bold=True),
        "LIGHT_GREEN": Font(name="微软雅黑", size=10, color="375623", bold=True),
        "MID_GREEN": Font(name="微软雅黑", size=10, color="375623", bold=True),
        "BLUE": Font(name="微软雅黑", size=10, color="1F4E79", bold=True),
        "LIGHT_RED": Font(name="微软雅黑", size=10, color="C65911"),
        "GRAY": Font(name="微软雅黑", size=10, color="7F7F7F"),
    }

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 美化表头
    for col_num in range(1, len(headers_order) + 1):
        cell = cast(Cell, ws.cell(row=1, column=col_num))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    cat_col_idx = headers_order.index("志愿类别") + 1
    trend_col_idx = headers_order.index("计划与位次趋势") + 1

    # 写入并美化数据
    for row_data in final_rows:
        row_values = [row_data.get(h, "") for h in headers_order]
        ws.append(row_values)
        curr_row = ws.max_row

        for col_num in range(1, len(headers_order) + 1):
            cell = cast(Cell, ws.cell(row=curr_row, column=col_num))
            cell.font = data_font
            cell.border = thin_border

            # 数字 and 文本的自适应对齐方式
            val_str = str(cell.value) if cell.value is not None else ""
            if val_str.isdigit():
                cell.value = int(val_str)  # 转换为数值型便于Excel公式计算
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # 针对“志愿类别”列作色彩渲染
            if col_num == cat_col_idx:
                cat_color = row_data.get("CAT_COLOR_TAG", "GRAY")
                cell.fill = fills.get(cat_color, fills["GRAY"])
                cell.font = text_colors.get(cat_color, text_colors["GRAY"])
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 针对“计划与位次趋势”列单独作高亮色彩渲染
            elif col_num == trend_col_idx:
                trend_color = row_data.get("COLOR_TAG", "GRAY")
                cell.fill = fills.get(trend_color, fills["GRAY"])
                cell.font = text_colors.get(trend_color, text_colors["GRAY"])
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自动调整列宽与冻结首行
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    for col in ws.columns:
        max_len = 0
        first_cell = col[0]
        col_idx = first_cell.column
        if col_idx is None:
            continue
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                # 兼容中文字符宽度计算
                cell_len = sum(
                    2 if "\u4e00" <= ch <= "\u9fff" else 1 for ch in str(cell.value)
                )
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

    # 保存结果
    output_path = os.path.join(OUTPUT_DIR, f"{base_filename}.xlsx")
    wb.save(output_path)
    print(
        f"[生成完毕] 成功创建最新的志愿辅助看板 ({len(final_rows)} 行数据): {output_path}"
    )


if __name__ == "__main__":
    generate_report()
